#!/usr/bin/env python3

import urllib
import urllib.request
import urllib.error
import json
import datetime
import sys
import csv
from openpyxl import Workbook

"""
make a call API-NG
"""

def callAping(jsonrpc_req):
    try:
        req = urllib.request.Request(url, jsonrpc_req.encode('utf-8'), headers)
        response = urllib.request.urlopen(req)
        jsonResponse = response.read()
        return jsonResponse.decode('utf-8')
    except urllib.error.URLError as e:
        print (e.reason) 
        print ('Oops no service available at ' + str(url))
        exit()
    except urllib.error.HTTPError:
        print ('Oops not a valid operation from the service ' + str(url))
        exit()


def callAping1(jsonrpc_req):
    try:
        req = urllib.request.Request(url1, jsonrpc_req.encode('utf-8'), headers)
        response = urllib.request.urlopen(req)
        jsonResponse = response.read()
        return jsonResponse.decode('utf-8')
    except urllib.error.URLError as e:
        print (e.reason) 
        print ('Oops no service available at ' + str(url1))
        exit()
    except urllib.error.HTTPError:
        print ('Oops not a valid operation from the service ' + str(url1))
        exit()


"""
calling getEventTypes operation
"""

def getEventTypes():
    event_type_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listEventTypes", "params": {"filter":{ }}, "id": 1}'
    print ('Calling listEventTypes to get event Type ID')
    eventTypesResponse = callAping(event_type_req)
    eventTypeLoads = json.loads(eventTypesResponse)
    #print(eventTypesResponse)
    try:
        eventTypeResults = eventTypeLoads['result']
        return eventTypeResults
    except:
        print ('Exception from API-NG' + str(eventTypeLoads['error']))
        exit()


"""
calling getEventTypes operation
"""
def getEvents(eventTypeID):
    today_from = datetime.datetime.now().strftime('%Y-%m-%d') + "T00:00:00"
    today_end = datetime.datetime.now().strftime('%Y-%m-%d') + "T23:59:59"
    event_type_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listEvents", "params": {"filter":{ "eventTypeIds":["' + eventTypeID + '"], "marketStartTime":{"from":"' + today_from + '","to":"' + today_end + \
        '"} }}, "id": 1}'
    print('Calling getEvents to get event Type ID')
    eventTypesResponse = callAping(event_type_req)
    eventTypeLoads = json.loads(eventTypesResponse)
    try:
        eventTypeResults = eventTypeLoads['result']
        return eventTypeResults
    except:
        print ('Exception from API-NG' + str(eventTypeLoads['error']))
        exit()

def getRaceDetails(meetingId):
    # "eventTypeIds":["' + meetingId + '"]
    event_type_req = '{"jsonrpc": "2.0", "method": "ScoresAPING/v1.0/listRaceDetails", "params": {"filter":{ "meetingIds": ["' + meetingId + '"] }}, "id": 1}'
    print ('Calling listRaceDetails')
    eventTypesResponse = callAping1(event_type_req)
    eventTypeLoads = json.loads(eventTypesResponse)
    print(eventTypesResponse)
    try:
        eventTypeResults = eventTypeLoads['result']
        return eventTypeResults
    except:
        print ('Exception from API-NG' + str(eventTypeLoads['error']))
        exit()

"""
calling getEventTypes operation
"""
def getCountries():
    event_type_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listCountries", "params": {"filter":{''}}, "id": 1}'
    print ('Calling listCountries')
    eventTypesResponse = callAping(event_type_req)
    eventTypeLoads = json.loads(eventTypesResponse)
    print(eventTypesResponse)
    try:
        eventTypeResults = eventTypeLoads['result']
        return eventTypeResults
    except:
        print ('Exception from API-NG' + str(eventTypeLoads['error']))
        exit()

def getMarketTypes():
    event_type_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listMarketTypes", "params": {"filter":{''}}, "id": 1}'
    print ('Calling listMarketTypes')
    eventTypesResponse = callAping(event_type_req)
    eventTypeLoads = json.loads(eventTypesResponse)
    print(eventTypesResponse)
    try:
        eventTypeResults = eventTypeLoads['result']
        return eventTypeResults
    except:
        print ('Exception from API-NG' + str(eventTypeLoads['error']))
        exit()


"""
Extraction eventypeId for eventTypeName from evetypeResults
"""

def getEventTypeIDForEventTypeName(eventTypesResult, requestedEventTypeName):
    if(eventTypesResult is not None):
        for event in eventTypesResult:
            eventTypeName = event['eventType']['name']
            if( eventTypeName == requestedEventTypeName):
                return event['eventType']['id']
    else:
        print ('Oops there is an issue with the input')
        exit()


"""
Calling marketCatalouge to get marketDetails
"""

def getMarketCatalogueForNextGBWin(eventTypeID, eventID):
    if (eventTypeID is not None):
        print ('Calling listMarketCatalouge Operation to get MarketID and selectionId')
        today_from = datetime.datetime.now().strftime('%Y-%m-%d') + "T00:00:00"
        today_end = datetime.datetime.now().strftime('%Y-%m-%d') + "T23:59:59"

        market_catalogue_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listMarketCatalogue", "params": {"filter":{"eventTypeIds":["' + eventTypeID + '"],"eventIds": ["' + eventID + '"], "marketTypeCodes":["WIN"],'\
                                                                                                            '"marketStartTime":{"from":"' + today_from + '","to":"' + today_end + \
                                                                                                            '"}},"sort":"FIRST_TO_START","maxResults":"1","marketProjection":["RUNNER_METADATA"]}, "id": 1}'
        #market_catalogue_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listMarketCatalogue", "params": {"filter":{"eventTypeIds":["' + eventTypeID + '"],"marketTypeCodes":["FORECAST"],'\
        #                                                                                                                                                     '"marketStartTime":{"from":"' + now + '"}},"sort":"FIRST_TO_START","maxResults":"10","marketProjection":["RUNNER_METADATA"]}, "id": 1}'
        """
        print(market_catalogue_req)
        """
        market_catalogue_response = callAping(market_catalogue_req)
        market_catalouge_loads = json.loads(market_catalogue_response)
        try:
            market_catalouge_results = market_catalouge_loads['result']
            return market_catalouge_results
        except:
            print ('Exception from API-NG' + str(market_catalouge_results['error']))
            exit()


def getMarketId(marketCatalogueResult):
    if( marketCatalogueResult is not None):
        for market in marketCatalogueResult:
            return market['marketId']


def getSelectionId(marketCatalogueResult):
    if(marketCatalogueResult is not None):
        for market in marketCatalogueResult:
            return market['runners'][0]['selectionId']


def getMarketBookBestOffers(marketId):
    print ('Calling listMarketBook to read prices for the Market with ID :' + marketId)
    market_book_req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/listMarketBook", "params": {"marketIds":["' + marketId + '"],"priceProjection":{"priceData":["EX_BEST_OFFERS"]}}, "id": 1}'
    """
    print(market_book_req)
    """
    market_book_response = callAping(market_book_req)
    """
    print(market_book_response)
    """
    market_book_loads = json.loads(market_book_response)
    try:
        market_book_result = market_book_loads['result']
        return market_book_result
    except:
        print ('Exception from API-NG' + str(market_book_result['error']))
        exit()


def printPriceInfo(market_book_result):
    if(market_book_result is not None):
        print ('Please find Best three available prices for the runners')
        for marketBook in market_book_result:
            runners = marketBook['runners']
            for runner in runners:
                if (runner['status'] == 'ACTIVE'):
                    print ('Available to back price :' + str(runner['ex']['availableToBack']))
                    print ('Available to lay price :' + str(runner['ex']['availableToLay']))
                else:
                    print ('This runner is not active')


def placeFailingBet(marketId, selectionId):
    if( marketId is not None and selectionId is not None):
        print ('Calling placeOrder for marketId :' + marketId + ' with selection id :' + str(selectionId))
        place_order_Req = '{"jsonrpc": "2.0", "method": "SportsAPING/v1.0/placeOrders", "params": {"marketId":"' + marketId + '","instructions":'\
                                                                                                                              '[{"selectionId":"' + str(
            selectionId) + '","handicap":"0","side":"BACK","orderType":"LIMIT","limitOrder":{"size":"0.01","price":"1.50","persistenceType":"LAPSE"}}],"customerRef":"test12121212121"}, "id": 1}'
        """
print(place_order_Req)
"""
        place_order_Response = callAping(place_order_Req)
        place_order_load = json.loads(place_order_Response)
        try:
            place_order_result = place_order_load['result']
            print ('Place order status is ' + place_order_result['status'])
            """
print('Place order error status is ' + place_order_result['errorCode'])
"""
            print ('Reason for Place order failure is ' + place_order_result['instructionReports'][0]['errorCode'])
        except:
            print ('Exception from API-NG' + str(place_order_result['error']))
        """
print(place_order_Response)
"""


url = "https://api.betfair.com/exchange/betting/json-rpc/v1"
url1 = "https://api.betfair.com/exchange/scores/json-rpc/v1"

"""
headers = { 'X-Application' : 'xxxxxx', 'X-Authentication' : 'xxxxx' ,'content-type' : 'application/json' }
"""

args = len(sys.argv)

appKey = "DZropQlOy2J5S7N2"
sessionToken = "tdKSaVZWkzhr+A+c27PnrtH16mesGiYmCCBA7Pl3ags="
sessionToken = "hYw0buwK4WzlYe2rmVB4/6rgJdupFKOmeEwhJ4H8TBU="

headers = {'X-Application': appKey, 'X-Authentication': sessionToken, 'content-type': 'application/json'}

#getCountries()
#getMarketTypes()

eventTypesResult = getEventTypes()
horseRacingEventTypeID = getEventTypeIDForEventTypeName(eventTypesResult, 'Horse Racing')

print ('Eventype Id for Horse Racing is :' + str(horseRacingEventTypeID))
events = getEvents(horseRacingEventTypeID)
#event_id = events[0]['event']['id'];
#print(event_id)

#getRaceDetails(event_id)

wb = Workbook()
sheet = wb.active
row = 1
for i in range(1, len(events)) :
    event = events[i]
    event_id = event['event']['id']
    name = event['event']['name']
    try:
        countryCode = event['event']['countryCode']
    except:
        countryCode = ""
    timezone = event['event']['timezone']
    openDate = event['event']['openDate']
    try:
        venue = event['event']['venue']
    except:
        venue = ""
    market_count = event['marketCount']
    try:
        market = getMarketCatalogueForNextGBWin(horseRacingEventTypeID, event_id)[0]
    except:
        continue
    market_id = ""
    marketInfo = {}
    if(market is not None):
        market_id = str(market["marketId"])
        marketInfo['marketid'] = str(market_id)
        marketInfo['runners'] = {}
        for runner in market['runners']:
            runnerInfo = {}
            runnerId = runner['selectionId']
            runnerInfo['runnerId'] = runnerId
            runnerInfo['runnerName'] = runner['runnerName']
            runnerInfo['max_back_odd'] = '0'
            runnerInfo['min_lay_odd'] = '0'
            marketInfo['runners'][runnerId] = runnerInfo

        market_book_result = getMarketBookBestOffers(market_id)
        if(market_book_result is not None):
            for marketBook in market_book_result:
                runners = marketBook['runners']
                for runner in runners:
                    print(runner)
                    runnerInfo = marketInfo['runners'][runner['selectionId']]
                    #if (runner['status'] == 'ACTIVE'):
                    try:
                        runnerInfo['max_back_odd'] = str(runner['ex']['availableToBack'][0]['price'])
                    except:
                        runnerInfo['max_back_odd'] = "0"
                    try:
                        runnerInfo['min_lay_odd'] = str(runner['ex']['availableToLay'][0]['price'])
                    except:
                        runnerInfo['min_lay_odd'] = "0"
                    #else:
                    #    print ('This runner is not active')

    sheet.cell(row=row, column=1).value = i
    sheet.cell(row=row, column=2).value = market_id
    sheet.cell(row=row, column=3).value = event_id
    sheet.cell(row=row, column=4).value = openDate
    sheet.cell(row=row, column=5).value = venue
    print(marketInfo['runners'])
    index = 6
    for runner in marketInfo['runners'] :
        print(runner)
        sheet.cell(row=row, column=6).value = marketInfo['runners'][runner]['runnerName']
        index = index + 1
        sheet.cell(row=row, column=index).value = marketInfo['runners'][runner]['runnerId']
        index = index + 1
        sheet.cell(row=row, column=index).value = marketInfo['runners'][runner]['max_back_odd']
        index = index + 1
        sheet.cell(row=row, column=index).value = marketInfo['runners'][runner]['min_lay_odd']
        index = index + 1
    row = row + 1

wb.save("horseracing.xlsx")

# info = []
# if(marketCatalogueResult is not None):
#     for market in marketCatalogueResult:
#         marketInfo = {}
#         marketid = market['marketId']
#         marketInfo['marketid'] = str(marketid)
#         marketInfo['runners'] = {}
#         for runner in market['runners']:
#             runnerInfo = {}
#             runnerId = runner['selectionId']
#             runnerInfo['runnerId'] = runnerId
#             runnerInfo['runnerName'] = runner['runnerName']
#             runnerInfo['max_back_odd'] = '0'
#             runnerInfo['min_lay_odd'] = '0'
#             marketInfo['runners'][runnerId] = runnerInfo
#         info.append(marketInfo)
#         market_book_result = getMarketBookBestOffers(marketid)
#         if(market_book_result is not None):
#             for marketBook in market_book_result:
#                 runners = marketBook['runners']
#                 for runner in runners:
#                     runnerInfo = marketInfo['runners'][runnerId]
#                     print ('Selection id is ' + str(runner['selectionId']))
#                     if (runner['status'] == 'ACTIVE'):
#                         runnerInfo['max_back_odd'] = str(runner['ex']['availableToBack'][0]['price'])
#                         runnerInfo['min_lay_odd'] = str(runner['ex']['availableToLay'][0]['price'])
#                     else:
#                         print ('This runner is not active')
#         venue_result = getVenues(marketid)
#         print(venue_result)
#         #printPriceInfo(market_book_result)
# print(info)

# marketid = getMarketId(marketCatalogueResult)
# runnerId = getSelectionId(marketCatalogueResult)

# print(marketid)
# print(runnerId)

# market_book_result = getMarketBookBestOffers(marketid)
# printPriceInfo(market_book_result)

#placeFailingBet(marketid, runnerId)
